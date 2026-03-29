# /app/backend/routes/reports_export.py
# تصدير التقارير بصيغة PDF و Excel

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta, timezone
from io import BytesIO
import os

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# MongoDB
from motor.motor_asyncio import AsyncIOMotorClient

router = APIRouter(prefix="/reports", tags=["Reports Export"])

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "trend_syria")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]


def format_price(price):
    """تنسيق السعر بالليرة السورية"""
    return f"{price:,.0f} ل.س"


def format_date(date):
    """تنسيق التاريخ"""
    if isinstance(date, datetime):
        return date.strftime("%Y-%m-%d %H:%M")
    return str(date)


async def get_sales_data(seller_id: str = None, start_date: datetime = None, end_date: datetime = None):
    """جلب بيانات المبيعات"""
    query = {"status": {"$in": ["paid", "delivered", "shipped"]}}
    
    if seller_id:
        query["seller_id"] = seller_id
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return orders


async def get_products_data(seller_id: str = None):
    """جلب بيانات المنتجات"""
    query = {}
    if seller_id:
        query["seller_id"] = seller_id
    
    products = await db.products.find(query, {"_id": 0}).to_list(1000)
    return products


@router.get("/sales/excel")
async def export_sales_excel(
    seller_id: str = Query(None, description="معرف البائع (اختياري)"),
    days: int = Query(30, description="عدد الأيام")
):
    """
    تصدير تقرير المبيعات بصيغة Excel
    """
    try:
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=days)
        
        orders = await get_sales_data(seller_id, start_date, end_date)
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المبيعات"
        ws.sheet_view.rightToLeft = True  # RTL for Arabic
        
        # تنسيق الهيدر
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # رؤوس الأعمدة
        headers = ["رقم الطلب", "التاريخ", "العميل", "المنتجات", "المبلغ", "الحالة", "طريقة الدفع"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # البيانات
        status_map = {
            "pending": "قيد الانتظار",
            "paid": "مدفوع",
            "shipped": "تم الشحن",
            "delivered": "تم التوصيل",
            "cancelled": "ملغي"
        }
        
        payment_map = {
            "cod": "الدفع عند الاستلام",
            "wallet": "المحفظة",
            "card": "بطاقة"
        }
        
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.get("order_number", "N/A"))
            ws.cell(row=row, column=2, value=format_date(order.get("created_at", "")))
            ws.cell(row=row, column=3, value=order.get("customer_name", "غير معروف"))
            
            # المنتجات
            items = order.get("items", [])
            products_str = ", ".join([f"{item.get('name', '')} x{item.get('quantity', 1)}" for item in items[:3]])
            if len(items) > 3:
                products_str += f" و{len(items)-3} منتج آخر"
            ws.cell(row=row, column=4, value=products_str)
            
            ws.cell(row=row, column=5, value=order.get("total", 0))
            ws.cell(row=row, column=6, value=status_map.get(order.get("status", ""), order.get("status", "")))
            ws.cell(row=row, column=7, value=payment_map.get(order.get("payment_method", ""), order.get("payment_method", "")))
        
        # تعديل عرض الأعمدة
        column_widths = [15, 20, 20, 40, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # إضافة ملخص
        summary_row = len(orders) + 3
        ws.cell(row=summary_row, column=1, value="الإجمالي:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        total_sales = sum(order.get("total", 0) for order in orders)
        ws.cell(row=summary_row, column=5, value=total_sales)
        ws.cell(row=summary_row, column=5).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value=f"عدد الطلبات: {len(orders)}")
        ws.cell(row=summary_row + 2, column=1, value=f"الفترة: {days} يوم")
        
        # حفظ في buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل إنشاء التقرير: {str(e)}")


@router.get("/sales/pdf")
async def export_sales_pdf(
    seller_id: str = Query(None, description="معرف البائع (اختياري)"),
    days: int = Query(30, description="عدد الأيام")
):
    """
    تصدير تقرير المبيعات بصيغة PDF
    """
    try:
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=days)
        
        orders = await get_sales_data(seller_id, start_date, end_date)
        
        # إنشاء PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # العنوان
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#FF6B00'),
            alignment=1  # Center
        )
        elements.append(Paragraph("تقرير المبيعات - ترند سوريا", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # معلومات التقرير
        info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, alignment=1)
        elements.append(Paragraph(f"الفترة: آخر {days} يوم", info_style))
        elements.append(Paragraph(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}", info_style))
        elements.append(Paragraph(f"عدد الطلبات: {len(orders)}", info_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # جدول المبيعات
        if orders:
            status_map = {
                "pending": "انتظار",
                "paid": "مدفوع",
                "shipped": "شحن",
                "delivered": "تسليم",
                "cancelled": "ملغي"
            }
            
            table_data = [["#", "رقم الطلب", "التاريخ", "المبلغ", "الحالة"]]
            
            for i, order in enumerate(orders[:50], 1):  # أول 50 طلب فقط
                table_data.append([
                    str(i),
                    order.get("order_number", "N/A"),
                    format_date(order.get("created_at", ""))[:10],
                    format_price(order.get("total", 0)),
                    status_map.get(order.get("status", ""), order.get("status", ""))
                ])
            
            # إضافة صف الإجمالي
            total_sales = sum(order.get("total", 0) for order in orders)
            table_data.append(["", "", "الإجمالي:", format_price(total_sales), ""])
            
            table = Table(table_data, colWidths=[1*cm, 3*cm, 3*cm, 4*cm, 2.5*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B00')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF5E6')),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')])
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("لا توجد طلبات في هذه الفترة", info_style))
        
        doc.build(elements)
        output.seek(0)
        
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل إنشاء التقرير: {str(e)}")


@router.get("/products/excel")
async def export_products_excel(
    seller_id: str = Query(None, description="معرف البائع (اختياري)")
):
    """
    تصدير تقرير المنتجات بصيغة Excel
    """
    try:
        products = await get_products_data(seller_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المنتجات"
        ws.sheet_view.rightToLeft = True
        
        # تنسيق الهيدر
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        headers = ["اسم المنتج", "الفئة", "السعر", "المخزون", "حالة الموافقة", "عدد المبيعات"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        approval_map = {
            "pending": "معلق",
            "approved": "موافق",
            "rejected": "مرفوض"
        }
        
        for row, product in enumerate(products, 2):
            ws.cell(row=row, column=1, value=product.get("name", ""))
            ws.cell(row=row, column=2, value=product.get("category", ""))
            ws.cell(row=row, column=3, value=product.get("price", 0))
            ws.cell(row=row, column=4, value=product.get("stock", 0))
            ws.cell(row=row, column=5, value=approval_map.get(product.get("approval_status", ""), ""))
            ws.cell(row=row, column=6, value=product.get("sales_count", 0))
        
        # عرض الأعمدة
        column_widths = [35, 15, 15, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # ملخص
        summary_row = len(products) + 3
        ws.cell(row=summary_row, column=1, value=f"إجمالي المنتجات: {len(products)}")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        total_stock = sum(p.get("stock", 0) for p in products)
        ws.cell(row=summary_row + 1, column=1, value=f"إجمالي المخزون: {total_stock}")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"products_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل إنشاء التقرير: {str(e)}")


@router.get("/analytics/excel")
async def export_analytics_excel(
    seller_id: str = Query(None, description="معرف البائع"),
    days: int = Query(30, description="عدد الأيام")
):
    """
    تصدير تقرير تحليلي شامل بصيغة Excel
    """
    try:
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=days)
        
        orders = await get_sales_data(seller_id, start_date, end_date)
        products = await get_products_data(seller_id)
        
        wb = Workbook()
        
        # ورقة الملخص
        ws_summary = wb.active
        ws_summary.title = "ملخص"
        ws_summary.sheet_view.rightToLeft = True
        
        header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # إحصائيات عامة
        total_sales = sum(order.get("total", 0) for order in orders)
        total_orders = len(orders)
        avg_order = total_sales / total_orders if total_orders > 0 else 0
        
        summary_data = [
            ["التقرير التحليلي الشامل", ""],
            ["", ""],
            ["الفترة", f"آخر {days} يوم"],
            ["تاريخ التصدير", datetime.now().strftime('%Y-%m-%d %H:%M')],
            ["", ""],
            ["إحصائيات المبيعات", ""],
            ["إجمالي المبيعات", f"{total_sales:,.0f} ل.س"],
            ["عدد الطلبات", total_orders],
            ["متوسط قيمة الطلب", f"{avg_order:,.0f} ل.س"],
            ["", ""],
            ["إحصائيات المنتجات", ""],
            ["عدد المنتجات", len(products)],
            ["إجمالي المخزون", sum(p.get("stock", 0) for p in products)],
        ]
        
        for row, data in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=data[0])
            ws_summary.cell(row=row, column=2, value=data[1])
            if row in [1, 6, 11]:  # العناوين
                ws_summary.cell(row=row, column=1).font = Font(bold=True, size=14)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 25
        
        # ورقة المبيعات اليومية
        ws_daily = wb.create_sheet("المبيعات اليومية")
        ws_daily.sheet_view.rightToLeft = True
        
        # تجميع المبيعات حسب اليوم
        daily_sales = {}
        for order in orders:
            date = order.get("created_at")
            if isinstance(date, datetime):
                day_key = date.strftime("%Y-%m-%d")
                if day_key not in daily_sales:
                    daily_sales[day_key] = {"count": 0, "total": 0}
                daily_sales[day_key]["count"] += 1
                daily_sales[day_key]["total"] += order.get("total", 0)
        
        headers = ["التاريخ", "عدد الطلبات", "إجمالي المبيعات"]
        for col, header in enumerate(headers, 1):
            cell = ws_daily.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, (date, data) in enumerate(sorted(daily_sales.items(), reverse=True), 2):
            ws_daily.cell(row=row, column=1, value=date)
            ws_daily.cell(row=row, column=2, value=data["count"])
            ws_daily.cell(row=row, column=3, value=data["total"])
        
        for i, width in enumerate([15, 15, 20], 1):
            ws_daily.column_dimensions[get_column_letter(i)].width = width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"analytics_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل إنشاء التقرير: {str(e)}")
